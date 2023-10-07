import openpyxl as openpyxl

def read_excel(file_ame):
    # 加载excel
    xls = openpyxl.load_workbook(file_ame)
    # 读取sheet
    sheet1 = xls['Sheet1']
    max_row = sheet1.max_row
    max_col = sheet1.max_column
    data_list = []
    for row in range(2, max_row + 1):
        row_list = []
        for col in range(1, max_col + 1):
            row_list.append(sheet1.cell(row, col).value)
        data_list.append(row_list)

    return data_list
